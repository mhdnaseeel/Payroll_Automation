import openpyxl
import pandas as pd
import os

import sys

if len(sys.argv) > 1:
    file_path = sys.argv[1]
else:
    file_path = "FCI _Automation_ToolNOV2025.xlsm"

if not os.path.exists(file_path):
    print(f"Error: File not found at {file_path}")
    exit(1)

print(f"Loading {file_path}...")
try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
except Exception as e:
    print(f"Error loading workbook: {e}")
    exit(1)

print("\n--- SHEETS ---")
for sheet in wb.sheetnames:
    ws = wb[sheet]
    state = ws.sheet_state
    print(f"Sheet: '{sheet}' (State: {state})")
    
    # Print first 5 rows to understand structure
    print("  Header/First few rows:")
    data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 5: break
        # filter out None
        row_data = [str(c).strip() if c is not None else "" for c in row]
        # if row is mostly empty, skip or print truncated
        if any(row_data):
            print(f"    Row {i+1}: {row_data}")
    print("-" * 30)

print("\n--- NAMED RANGES ---")
for name in wb.defined_names.definedName:
    print(f"Name: {name.name}, RefersTo: {name.value}")

print("\n--- ANALYSIS COMPLETE ---")
